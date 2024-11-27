from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def format_text_to_width(text, max_width=50):
    """
    格式化文本，每行最多 max_width 个字符，自动换行。
    """
    words = text.split()
    lines = []
    current_line = ""
    for word in words:
        if len(current_line) + len(word) + 1 > max_width:
            lines.append(current_line)
            current_line = word
        else:
            current_line += (" " + word if current_line else word)
    if current_line:
        lines.append(current_line)
    return "\n".join(lines)

def extract_and_save_columns_to_txt(excel_file, output_txt_file, columns, max_line_length=50):
    """
    从 Excel 文件中提取指定列的内容，并保存为格式化的文本文件。
    
    :param excel_file: 输入的 Excel 文件路径
    :param output_txt_file: 输出的文本文件路径
    :param columns: 要提取的列的名称，例如 ["A", "C", "F"]
    :param max_line_length: 文本文档每行最大字符数
    """
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    with open(output_txt_file, "w", encoding="utf-8") as txt_file:
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始读取，跳过标题行
            extracted_values = [row[column_index_from_string(col) - 1] for col in columns]
            # 如果 F 列为空，跳过此行
            if not extracted_values[-1]:  # 检查最后一列是否为空
                continue
            
            # 格式化每个单元格内容，并添加空行
            formatted_lines = []
            for value in extracted_values:
                if value:
                    formatted_lines.append(format_text_to_width(str(value), max_width=max_line_length))
                formatted_lines.append("")  # 添加空行
            formatted_lines = formatted_lines[:-1]  # 移除最后多余的空行
            
            # 写入到文本文件
            txt_file.write("\n".join(formatted_lines))
            txt_file.write("\n" + "-" * 50 + "\n")  # 添加更长的分隔符

    print(f"提取完成，结果已保存到 {output_txt_file}")

# 示例用法
excel_file_path = r"\\ADMIN-20231023E\software\工作文件夹\项目代码\小说推文批量处理\Data\文本\总识别结果\总识别结果_模型整理文本_审核是否有性暗示\总识别结果_模型整理文本_审核是否有性暗示_模型转译.xlsx"  # 替换为您的 Excel 文件路径
output_txt_file_path = r"data\24-11-26汇总文档.txt"  # 替换为您想保存的文本文件路径
columns_to_extract = ["A", "C", "F"]  # 替换为您想提取的列名

extract_and_save_columns_to_txt(excel_file_path, output_txt_file_path, columns_to_extract)
